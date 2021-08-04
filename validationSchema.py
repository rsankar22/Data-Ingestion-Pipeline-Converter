#this is where the magic happens
from fileInScan import pullFile
from openpyxl import load_workbook

#let us start with some basic database schemas; ie xlsx, csv, dbo etc.
#BEGIN WITH EXCEL STUFF
def excel_valid(file):
    workbook = load_workbook(file)
    return workbook
    sheet = workbook.get_sheet_by_name('Sheet1')
    for i in sheet:
        for j in sheet[i]:
            return  (sheet.cell(row=i,column=j).value)